import json
import time
import re
from openpyxl import load_workbook

# ====== C·∫§U H√åNH C∆† B·∫¢N ======
FILE_PATH = 'data.xlsx'
KEY_FILE_PATH = 'key.json'
LANG_SOURCE = 'vi'
LANG_TARGET = 'zh'          # 'zh' ho·∫∑c 'zh-CN' t√πy b·∫°n
OUTPUT_SUFFIX = "_Dich_CN"
LOCATION = 'asia-southeast1'  # 'us-central1' c≈©ng ƒë∆∞·ª£c

# ====== KH·ªûI T·∫†O VERTEX AI (GEMINI) ======
from google.oauth2 import service_account
import vertexai
from vertexai.generative_models import GenerativeModel

try:
    with open(KEY_FILE_PATH, 'r', encoding='utf-8') as f:
        sa_info = json.load(f)
    PROJECT_ID = sa_info.get("project_id")
    if not PROJECT_ID:
        raise ValueError("Kh√¥ng t√¨m th·∫•y 'project_id' trong t·ªáp service account.")

    credentials = service_account.Credentials.from_service_account_file(KEY_FILE_PATH)
    vertexai.init(project=PROJECT_ID, location=LOCATION, credentials=credentials)
    model = GenerativeModel("gemini-1.5-pro")
    print(f"‚úÖ ƒê√£ kh·ªüi t·∫°o Gemini (Vertex AI) cho project: {PROJECT_ID} t·∫°i {LOCATION}")
except Exception as e:
    print(f"‚ùå L·ªñI KH·ªûI T·∫†O GEMINI/Vertex AI: {e}")
    raise SystemExit(1)

# √Ånh x·∫° t√™n sheet g·ªëc -> sheet m·ªõi
SHEET_NAME_MAP = {}

# ====== H√ÄM D·ªäCH B·∫∞NG GEMINI ======
import json as _json

def translate_batch_gemini(texts, lang_src, lang_tgt, max_retries=3, sleep_seconds=2):
    """
    D·ªãch danh s√°ch 'texts' b·∫±ng Gemini. Tr·∫£ v·ªÅ list c√°c chu·ªói ƒë√£ d·ªãch, c√πng th·ª© t·ª±.
    ƒê·ªÉ gom batch an to√†n, ta y√™u c·∫ßu model tr·∫£ v·ªÅ JSON array theo th·ª© t·ª±.
    """
    if not texts:
        return []

    # Prompt y√™u c·∫ßu tr·∫£ JSON array, gi·ªØ th·ª© t·ª±, kh√¥ng th√™m ch√∫ th√≠ch
    system_instruction = (
        "B·∫°n l√† c√¥ng c·ª• d·ªãch thu·∫≠t. H√£y d·ªãch ch√≠nh x√°c, t·ª± nhi√™n.\n"
        f"Ng√¥n ng·ªØ ngu·ªìn: {lang_src}\n"
        f"Ng√¥n ng·ªØ ƒë√≠ch: {lang_tgt}\n"
        "- Ch·ªâ d·ªãch ph·∫ßn vƒÉn b·∫£n; n·∫øu l√† c√¥ng th·ª©c Excel ho·∫∑c tham chi·∫øu √¥/sheet th√¨ b·ªè qua (nh∆∞ng ·ªü ƒë√¢y ƒë·∫ßu v√†o ƒë√£ l·ªçc c√¥ng th·ª©c r·ªìi).\n"
        "- Gi·ªØ nguy√™n s·ªë, m√£ SKU, k√Ω hi·ªáu ƒë·∫∑c bi·ªát n·∫øu kh√¥ng c·∫ßn d·ªãch.\n"
        "- B·∫£o to√†n xu·ªëng d√≤ng v√† kho·∫£ng tr·∫Øng quan tr·ªçng.\n"
        "- Tr·∫£ K·∫æT QU·∫¢ DUY NH·∫§T l√† m·ªôt JSON array c√°c chu·ªói, theo ƒë√∫ng th·ª© t·ª± input, kh√¥ng th√™m gi·∫£i th√≠ch."
    )

    # ƒê·ªÉ ch·∫Øc ƒÉn m·∫£ng JSON, m√¨nh ƒë√≥ng g√≥i input th√†nh JSON array string
    payload = _json.dumps(texts, ensure_ascii=False)

    content = [
        {"role": "user", "parts": [
            f"{system_instruction}\n\nƒê√¢y l√† danh s√°ch c√¢u c·∫ßn d·ªãch (JSON array):\n{payload}\n\n"
            "H√£y tr·∫£ v·ªÅ *ch·ªâ* m·ªôt JSON array c√°c b·∫£n d·ªãch, theo th·ª© t·ª± t∆∞∆°ng ·ª©ng."
        ]}
    ]

    for attempt in range(1, max_retries + 1):
        try:
            resp = model.generate_content(content, generation_config={
                "temperature": 0.2,
                "max_output_tokens": 2048,
                # B·∫°n c√≥ th·ªÉ set th√™m "response_mime_type": "application/json" (SDK m·ªõi h·ªó tr·ª£)
            })
            text = resp.text.strip()
            # C·ªë g·∫Øng parse JSON tr·ª±c ti·∫øp
            # M·ªôt s·ªë tr∆∞·ªùng h·ª£p model c√≥ th·ªÉ bao code block, ta l√†m s·∫°ch nh·∫π
            if text.startswith("```"):
                # lo·∫°i b·ªè ```json ... ```
                text = re.sub(r"^```(?:json)?\s*|\s*```$", "", text, flags=re.DOTALL)

            out = _json.loads(text)
            if not isinstance(out, list):
                raise ValueError("Ph·∫£n h·ªìi kh√¥ng ph·∫£i JSON array.")
            # ƒê·∫£m b·∫£o ƒë·ªô d√†i kh·ªõp:
            if len(out) != len(texts):
                raise ValueError(f"S·ªë ph·∫ßn t·ª≠ ph·∫£n h·ªìi ({len(out)}) kh√¥ng kh·ªõp input ({len(texts)}).")
            # Chu·∫©n h√≥a v·ªÅ str
            return [str(x) if x is not None else "" for x in out]

        except Exception as e:
            print(f"‚ö†Ô∏è L·ªói d·ªãch batch (l·∫ßn {attempt}/{max_retries}): {e}")
            if attempt < max_retries:
                time.sleep(sleep_seconds)
            else:
                # Th·ª≠ fallback: d·ªãch t·ª´ng c√¢u ƒë·ªÉ kh√¥ng m·∫•t k·∫øt qu·∫£ to√†n batch
                print("üîÅ Fallback: d·ªãch t·ª´ng c√¢u‚Ä¶")
                results = []
                for t in texts:
                    try:
                        single = translate_batch_gemini([t], lang_src, lang_tgt, max_retries=1, sleep_seconds=1)
                        results.append(single[0] if single else "")
                    except Exception:
                        results.append("")
                return results

def update_formula_references(sheet, sheet_map, output_suffix):
    """C·∫≠p nh·∫≠t tham chi·∫øu c√¥ng th·ª©c trong sheet ƒë·ªÉ tr·ªè ƒë·∫øn sheet ƒë√≠ch m·ªõi."""
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.data_type == 'f' and cell.value:
                formula = cell.value
                for old_name, new_name in sheet_map.items():
                    # kh·ªõp c·∫£ 'Sheet Name'!A1 v√† SheetName!A1
                    pattern = re.compile(r"([']?)" + re.escape(old_name) + r"([']?)!")
                    if pattern.search(formula):
                        new_formula = pattern.sub(rf"\1{new_name}\2!", formula)
                        cell.value = new_formula
                        formula = new_formula

def translate_and_copy_sheet(workbook, source_sheet_name):
    """Duy·ªát √¥ text, d·ªãch b·∫±ng Gemini, v√† ghi v√†o sheet m·ªõi (b·∫£n sao gi·ªØ ƒë·ªãnh d·∫°ng + c√¥ng th·ª©c)."""
    source_sheet = workbook[source_sheet_name]
    new_sheet_name = f"{source_sheet_name}{OUTPUT_SUFFIX}"
    SHEET_NAME_MAP[source_sheet_name] = new_sheet_name

    if new_sheet_name in workbook.sheetnames:
        print(f"‚ö†Ô∏è Sheet ƒë√≠ch '{new_sheet_name}' ƒë√£ t·ªìn t·∫°i. D·ªØ li·ªáu s·∫Ω ƒë∆∞·ª£c GHI ƒê√à.")
        new_sheet = workbook[new_sheet_name]
    else:
        new_sheet = workbook.copy_worksheet(source_sheet)
        new_sheet.title = new_sheet_name

    print(f"\n--- B·∫Øt ƒë·∫ßu d·ªãch Sheet: {source_sheet_name} ---")

    # Thu th·∫≠p c√°c √¥ text (kh√¥ng ph·∫£i c√¥ng th·ª©c)
    cells_to_translate = []
    for row in source_sheet.iter_rows(values_only=False):
        for cell in row:
            if (isinstance(cell.value, str) and cell.value.strip() and cell.data_type != 'f'):
                cells_to_translate.append({
                    'text': cell.value,
                    'row': cell.row,
                    'col': cell.column
                })

    original_texts = [item['text'] for item in cells_to_translate]

    # B·∫°n c√≥ th·ªÉ chia batch nh·ªè ƒë·ªÉ ·ªïn ƒë·ªãnh h∆°n (v√≠ d·ª• 100 m·ª•c/l∆∞·ª£t)
    BATCH_SIZE = 100
    translated_texts_all = []
    try:
        for i in range(0, len(original_texts), BATCH_SIZE):
            batch = original_texts[i:i+BATCH_SIZE]
            translated_batch = translate_batch_gemini(batch, LANG_SOURCE, LANG_TARGET)
            translated_texts_all.extend(translated_batch)
        print(f"‚úÖ ƒê√£ d·ªãch th√†nh c√¥ng {len(translated_texts_all)} ƒëo·∫°n text b·∫±ng Gemini.")
    except Exception as e:
        print(f"‚ùå L·ªñI D·ªäCH GEMINI: {e}")
        return

    # Ghi ƒë√® k·∫øt qu·∫£ v√†o sheet m·ªõi
    for i, item in enumerate(cells_to_translate):
        target_cell = new_sheet.cell(row=item['row'], column=item['col'])
        target_cell.value = translated_texts_all[i]

    print(f"--- Ho√†n th√†nh ghi ƒë√® Sheet: {source_sheet_name} ---")

# ====== CH∆Ø∆†NG TR√åNH CH√çNH ======
try:
    wb = load_workbook(FILE_PATH)
    original_sheet_names = [name for name in wb.sheetnames if OUTPUT_SUFFIX not in name]

    # B∆∞·ªõc 1: D·ªãch & t·∫°o sheet m·ªõi
    for sheet_name in original_sheet_names:
        translate_and_copy_sheet(wb, sheet_name)

    # B∆∞·ªõc 2: C·∫≠p nh·∫≠t tham chi·∫øu c√¥ng th·ª©c gi·ªØa c√°c sheet
    print("\n[B∆Ø·ªöC 2] B·∫Øt ƒë·∫ßu c·∫≠p nh·∫≠t tham chi·∫øu c√¥ng th·ª©c‚Ä¶")
    for sheet_name in original_sheet_names:
        new_sheet_name = SHEET_NAME_MAP[sheet_name]
        new_sheet = wb[new_sheet_name]
        update_formula_references(new_sheet, SHEET_NAME_MAP, OUTPUT_SUFFIX)
        print(f"‚úÖ ƒê√£ c·∫≠p nh·∫≠t tham chi·∫øu cho sheet: {new_sheet_name}")

    # B∆∞·ªõc 3: L∆∞u file m·ªõi
    OUTPUT_FILE_PATH = FILE_PATH.replace(".xlsx", "_DICH_CN.xlsx")
    wb.save(OUTPUT_FILE_PATH)
    print(f"\n‚úÖ TH√ÄNH C√îNG! File ƒë√£ d·ªãch ƒë∆∞·ª£c l∆∞u t·∫°i: {OUTPUT_FILE_PATH}")

except Exception as e:
    print(f"‚ùå L·ªói t·ªïng qu√°t khi x·ª≠ l√Ω file: {e}")
    print("H√£y ki·ªÉm tra xem file Excel c√≥ ƒëang m·ªü/kho√°, KEY_FILE_PATH ƒë√∫ng v√† service account c√≥ quy·ªÅn Vertex AI.")
